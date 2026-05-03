from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib.auth import authenticate, login
from django.contrib.auth.models import User
from django.db.models import Q, Count
from django.views.decorators.csrf import csrf_exempt
from .models import Food, Category, UnitConversion, SearchQuery
import json
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.conf import settings

# ========== SUPABASE HELPERS ==========
def get_all_foods_supabase():
    """Get all foods from Supabase"""
    try:
        response = settings.supabase.table('foods').select('*').execute()
        return response.data
    except Exception as e:
        print(f"Supabase error in get_all_foods_supabase: {e}")
        return []

def get_food_by_id_supabase(food_id):
    """Get a single food by ID from Supabase"""
    try:
        response = settings.supabase.table('foods').select('*').eq('id', food_id).execute()
        return response.data[0] if response.data else None
    except Exception as e:
        print(f"Supabase error in get_food_by_id_supabase: {e}")
        return None

def search_foods_supabase(query):
    """Search foods by name in Supabase"""
    try:
        response = settings.supabase.table('foods').select('*').ilike('food_name', f'%{query}%').limit(50).execute()
        return response.data
    except Exception as e:
        print(f"Supabase error in search_foods_supabase: {e}")
        return []

# ========== CONVERSION HELPERS ==========
def dict_to_food_object(data):
    """Convert Supabase dict to object with attributes for template compatibility"""
    class FoodItem:
        pass
    food = FoodItem()
    for key, value in data.items():
        setattr(food, key, value)
    return food

def dict_list_to_food_objects(data_list):
    """Convert list of Supabase dicts to list of objects"""
    return [dict_to_food_object(item) for item in data_list]

# ========== HELPER FUNCTIONS ==========
def get_cost_tag(food_name):
    """Determine if a food is affordable, medium, or high cost based on name"""
    food_name_lower = food_name.lower()
    
    affordable_keywords = ['maize', 'beans', 'sukuma', 'cabbage', 'dagaa', 'omena', 
                           'sweet potato', 'cassava', 'spinach', 'amaranth', 'millet',
                           'sorghum', 'githeri', 'mukimo', 'chapati', 'ugali']
    
    expensive_keywords = ['beef', 'lamb', 'pork', 'chicken', 'fish', 'milk', 'cheese',
                          'butter', 'ghee', 'yoghurt', 'sausage', 'pilau', 'biryani',
                          'samosa', 'cake', 'biscuit', 'bread', 'soda', 'juice']
    
    for kw in affordable_keywords:
        if kw in food_name_lower:
            return 'low'
    
    for kw in expensive_keywords:
        if kw in food_name_lower:
            return 'high'
    
    return 'medium'

# ========== SWAP SUGGESTIONS ==========
SWAP_SUGGESTIONS = {
    'maize': [
        {'name': 'whole maize flour', 'benefit': '3x more fiber and iron than refined maize'},
        {'name': 'sorghum', 'benefit': 'Higher iron content, good for anemia prevention'},
        {'name': 'finger millet (wimbi)', 'benefit': 'Rich in calcium and iron'}
    ],
    'rice': [
        {'name': 'brown rice', 'benefit': 'More fiber, B vitamins, and minerals'},
        {'name': 'sorghum', 'benefit': 'Higher iron and protein content'},
        {'name': 'finger millet', 'benefit': 'Excellent calcium source'}
    ],
    'beans': [
        {'name': 'green grams (ndengu)', 'benefit': 'Easier to digest, rich in iron'},
        {'name': 'lentils', 'benefit': 'Cook faster, high in folate and iron'},
        {'name': 'soybeans', 'benefit': 'Complete protein, high iron content'}
    ],
    'sukuma': [
        {'name': 'amaranth (terere)', 'benefit': 'Higher iron and calcium than sukuma'},
        {'name': 'spider plant (saget)', 'benefit': 'Rich in iron and protein'},
        {'name': 'kale (kanzera)', 'benefit': 'Similar nutrients, different taste'}
    ],
    'chapati': [
        {'name': 'whole wheat chapati', 'benefit': 'More fiber and minerals'},
        {'name': 'multigrain chapati', 'benefit': 'Added nutrients from millet and sorghum'},
        {'name': 'ugali with sukuma', 'benefit': 'Complete meal with more nutrients'}
    ],
    'beef': [
        {'name': 'fish (omena/dagaa)', 'benefit': 'Lower fat, rich in calcium and iron'},
        {'name': 'beans', 'benefit': 'Affordable plant protein with iron'},
        {'name': 'chicken without skin', 'benefit': 'Lower saturated fat'}
    ],
    'milk': [
        {'name': 'fermented milk (maziwa lala)', 'benefit': 'Easier digestion, probiotics'},
        {'name': 'fortified milk', 'benefit': 'Added vitamin A and D'},
        {'name': 'yoghurt', 'benefit': 'Probiotics, easier digestion'}
    ],
    'oil': [
        {'name': 'red palm oil', 'benefit': 'Rich in vitamin A'},
        {'name': 'olive oil', 'benefit': 'Heart-healthy monounsaturated fats'},
        {'name': 'use less oil', 'benefit': 'Reduce calories and fat intake'}
    ]
}

# ========== NUTRIENT SWAPS ==========
NUTRIENT_SWAPS = {
    'low_iron': [
        ('sukuma wiki', 'Excellent iron source, affordable'),
        ('beans', 'Rich in iron and protein'),
        ('dagaa omena', 'High iron and calcium'),
        ('amaranth (terere)', 'Very high iron content'),
        ('beef liver', 'Concentrated iron source')
    ],
    'low_fiber': [
        ('whole maize flour', '3x more fiber than refined'),
        ('beans', 'Excellent fiber source'),
        ('sweet potato', 'Good fiber content'),
        ('githeri', 'Mixed maize and beans for fiber')
    ],
    'low_calcium': [
        ('dagaa omena', 'Eaten with bones for calcium'),
        ('milk', 'Excellent calcium source'),
        ('amaranth (terere)', 'High calcium leafy green'),
        ('finger millet (wimbi)', 'Rich in calcium')
    ]
}

# ========== MAIN VIEWS ==========
def home(request):
    """Homepage with search"""
    popular_searches = SearchQuery.objects.all().order_by('-count')[:10]
    
    return render(request, 'foods/home.html', {
        'popular_searches': popular_searches,
    })

def search_foods(request):
    query = request.GET.get('q', '')
    category_id = request.GET.get('category', '')
    
    # Track search query
    if query:
        search_obj, created = SearchQuery.objects.get_or_create(query=query.lower())
        if not created:
            search_obj.count += 1
            search_obj.save()
    
    # Get foods from Supabase
    if query:
        foods_data = search_foods_supabase(query)
    else:
        foods_data = get_all_foods_supabase()
    
    # Convert to objects for template compatibility
    foods = dict_list_to_food_objects(foods_data)
    
    # Get categories
    categories = Category.objects.all().order_by('name')
    
    return render(request, 'foods/search_results.html', {
        'foods': foods,
        'query': query,
        'categories': categories,
        'selected_category': int(category_id) if category_id else None
    })

def food_detail(request, food_id):
    """Show detailed nutrient information for a specific food"""
    try:
        # Get food from Supabase
        food_data = get_food_by_id_supabase(food_id)
        
        if not food_data:
            return render(request, 'foods/error.html', {'error': 'Food not found'}, status=404)
        
        # Convert to object
        food = dict_to_food_object(food_data)

        # Track viewed foods
        if 'viewed_foods' not in request.session:
            request.session['viewed_foods'] = []
        
        viewed = request.session['viewed_foods']
        viewed.append(food.food_name)
        request.session['viewed_foods'] = viewed[-10:]
        
        # Calculate cost tag
        current_cost = get_cost_tag(food.food_name)
        
        # Generate swap suggestions
        swap_suggestions = []
        food_name_lower = food.food_name.lower()
        
        for keyword, swaps in SWAP_SUGGESTIONS.items():
            if keyword in food_name_lower:
                for swap in swaps:
                    if isinstance(swap, dict):
                        swap_name = swap.get('name', '')
                        benefit = swap.get('benefit', '')
                    elif isinstance(swap, (list, tuple)) and len(swap) >= 2:
                        swap_name = swap[0]
                        benefit = swap[1]
                    else:
                        continue
                        
                    try:
                        swap_food_data = search_foods_supabase(swap_name)
                        if swap_food_data:
                            swap_food = dict_to_food_object(swap_food_data[0])
                            if swap_food.id != food.id:
                                swap_suggestions.append({
                                    'name': swap_food.food_name,
                                    'id': swap_food.id,
                                    'benefit': benefit
                                })
                    except:
                        pass
        
        # Add affordable swaps if food is expensive
        if current_cost == 'high':
            affordable_swaps = [
                ('beans', 'Affordable plant protein — 1/4 the price of meat'),
                ('sukuma wiki', 'Iron-rich vegetable, very affordable'),
                ('dagaa omena', 'Calcium-rich fish, much cheaper than beef'),
                ('whole maize flour', 'Nutritious staple, budget-friendly'),
                ('cabbage', 'Vitamin-rich, very affordable'),
            ]
            for swap_name, benefit in affordable_swaps:
                try:
                    swap_food_data = search_foods_supabase(swap_name)
                    if swap_food_data:
                        swap_food = dict_to_food_object(swap_food_data[0])
                        if swap_food.id != food.id:
                            if not any(s['id'] == swap_food.id for s in swap_suggestions):
                                swap_suggestions.append({
                                    'name': swap_food.food_name,
                                    'id': swap_food.id,
                                    'benefit': f'💰 Affordable alternative — {benefit}'
                                })
                except:
                    pass
        
        # Add nutrient-based swaps if needed
        if len(swap_suggestions) < 3:
            if food.iron_mg < 2.0:
                for swap_name, benefit in NUTRIENT_SWAPS['low_iron']:
                    try:
                        swap_food_data = search_foods_supabase(swap_name)
                        if swap_food_data:
                            swap_food = dict_to_food_object(swap_food_data[0])
                            if swap_food.id != food.id:
                                if not any(s['id'] == swap_food.id for s in swap_suggestions):
                                    swap_suggestions.append({
                                        'name': swap_food.food_name,
                                        'id': swap_food.id,
                                        'benefit': f'🍳 High in iron — {benefit}'
                                    })
                    except:
                        pass
            
            if food.fiber_g < 3.0 and len(swap_suggestions) < 3:
                for swap_name, benefit in NUTRIENT_SWAPS['low_fiber']:
                    try:
                        swap_food_data = search_foods_supabase(swap_name)
                        if swap_food_data:
                            swap_food = dict_to_food_object(swap_food_data[0])
                            if swap_food.id != food.id:
                                if not any(s['id'] == swap_food.id for s in swap_suggestions):
                                    swap_suggestions.append({
                                        'name': swap_food.food_name,
                                        'id': swap_food.id,
                                        'benefit': f'🌾 High in fiber — {benefit}'
                                    })
                    except:
                        pass
        
        swap_suggestions = swap_suggestions[:3]
        
        return render(request, 'foods/food_detail.html', {
            'food': food,
            'swap_suggestions': swap_suggestions,
            'cost_tag': current_cost,
        })
        
    except Exception as e:
        print(f"Error in food_detail: {e}")
        return render(request, 'foods/error.html', {
            'error': str(e),
            'food_id': food_id
        }, status=500)

def test(request):
    """Simple test view to check if Django is running"""
    return HttpResponse("✅ Server is reachable! Django is working.")

def get_units(request):
    """Return available units for a food as JSON"""
    food_id = request.GET.get('food_id')
    if food_id:
        try:
            units = UnitConversion.objects.filter(food_id=food_id)
            unit_list = [{'name': u.unit_name, 'grams': u.grams} for u in units]
            return JsonResponse({'units': unit_list})
        except:
            return JsonResponse({'units': []})
    return JsonResponse({'units': []})

# ========== EXCEL EXPORTS ==========
def export_food_excel(request, food_id):
    """Export a single food's nutrient data as Excel file for nutritionists"""
    try:
        food_data = get_food_by_id_supabase(food_id)
        
        if not food_data:
            return HttpResponse(f"Food with ID {food_id} not found", status=404)
        
        food = dict_to_food_object(food_data)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{food.food_name[:25]} Nutrition"
        
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="2c5e2e", end_color="2c5e2e", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws['A1'] = "🇰🇪 Kenya Food Composition Database"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = food.food_name
        ws['A2'].font = Font(bold=True, size=12, italic=True)
        ws['A4'] = "Category:"
        ws['A4'].font = Font(bold=True)
        ws['B4'] = "Kenyan Food"
        
        ws['A6'] = "Nutrient"
        ws['B6'] = "Value per 100g"
        ws['C6'] = "Unit"
        
        for col in ['A6', 'B6', 'C6']:
            ws[col].font = header_font
            ws[col].fill = header_fill
            ws[col].alignment = header_alignment
            ws[col].border = border
        
        nutrients = [
            ('Energy', food.energy_kcal, 'kcal'),
            ('Protein', food.protein_g, 'g'),
            ('Total Fat', food.fat_g, 'g'),
            ('Carbohydrates', food.carbohydrate_g, 'g'),
            ('Dietary Fiber', food.fiber_g, 'g'),
            ('Iron', food.iron_mg, 'mg'),
            ('Calcium', food.calcium_mg, 'mg'),
        ]
        
        row = 7
        for name, value, unit in nutrients:
            formatted_value = '0' if value is None or value == 0 else f'{value:.1f}'
            ws[f'A{row}'] = name
            ws[f'B{row}'] = formatted_value
            ws[f'C{row}'] = unit
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = border
            row += 1
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"{food.food_name.replace(' ', '_')}_nutrition.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return HttpResponse(f"Error exporting data: {e}", status=500)

def export_recall_excel(request):
    """Export 24-hour recall results as Excel file for nutritionists"""
    try:
        if request.method == 'POST':
            from .models import Food, UnitConversion
            
            name = request.POST.get('name', '')
            age = request.POST.get('age', '')
            gender = request.POST.get('gender', 'female')
            
            try:
                age = int(age) if age else 30
            except ValueError:
                age = 30
            
            total_energy = 0
            total_protein = 0
            total_iron = 0
            total_calcium = 0
            total_fluid_ml = 0
            
            meals = [
                {'id': 1, 'name': 'Breakfast'},
                {'id': 2, 'name': 'Morning Snack'},
                {'id': 3, 'name': 'Lunch'},
                {'id': 4, 'name': 'Afternoon Snack'},
                {'id': 5, 'name': 'Dinner'},
                {'id': 6, 'name': 'Evening Snack'},
            ]
            
            food_items = []
            
            for meal in meals:
                meal_id = meal['id']
                food_ids = request.POST.getlist(f'food_id_{meal_id}[]')
                amounts = request.POST.getlist(f'amount_{meal_id}[]')
                units = request.POST.getlist(f'unit_{meal_id}[]')
                
                for i in range(len(food_ids)):
                    if food_ids[i] and amounts[i]:
                        try:
                            food_data = get_food_by_id_supabase(int(food_ids[i]))
                            if food_data:
                                food = dict_to_food_object(food_data)
                                amount = float(amounts[i])
                                unit = units[i] if i < len(units) else 'grams'
                                
                                grams = amount
                                if unit != 'grams':
                                    try:
                                        conversion = UnitConversion.objects.get(food_id=food_ids[i], unit_name=unit)
                                        grams = amount * conversion.grams
                                    except:
                                        pass
                                
                                energy = (grams / 100) * food.energy_kcal
                                protein = (grams / 100) * food.protein_g
                                iron = (grams / 100) * food.iron_mg
                                calcium = (grams / 100) * food.calcium_mg
                                
                                total_energy += energy
                                total_protein += protein
                                total_iron += iron
                                total_calcium += calcium
                                total_fluid_ml += (grams / 100) * (getattr(food, 'water_g', 0))
                                
                                food_items.append({
                                    'meal': meal['name'],
                                    'food': food.food_name,
                                    'amount': amount,
                                    'unit': unit,
                                    'grams': grams,
                                    'energy': energy,
                                    'protein': protein,
                                    'iron': iron,
                                    'calcium': calcium
                                })
                        except Exception as e:
                            print(f"Error processing food: {e}")
                            pass
            
            # Process fluids
            fluid_ids = request.POST.getlist('fluid_id[]')
            fluid_amounts = request.POST.getlist('fluid_amount[]')
            fluid_units = request.POST.getlist('fluid_unit[]')
            
            fluid_items = []
            for i in range(len(fluid_ids)):
                if fluid_ids[i] and fluid_amounts[i]:
                    try:
                        amount = float(fluid_amounts[i])
                        unit = fluid_units[i]
                        
                        ml = amount
                        if unit == 'cup':
                            ml = amount * 240
                        elif unit == 'glass':
                            ml = amount * 250
                        elif unit == 'bottle':
                            ml = amount * 500
                        else:
                            ml = amount
                        
                        total_fluid_ml += ml
                        
                        fluid_name = "Water"
                        if fluid_ids[i] != '9991':
                            fluid_map = {'9992': 'Black Tea', '9993': 'Milk Tea', '9994': 'Coffee',
                                        '9995': 'Orange Juice', '9996': 'Soda', '9997': 'Fermented Milk',
                                        '9998': 'Yogurt Drink', '9999': 'Fresh Juice'}
                            fluid_name = fluid_map.get(fluid_ids[i], 'Beverage')
                        
                        fluid_items.append({
                            'fluid': fluid_name,
                            'amount': amount,
                            'unit': unit,
                            'ml': ml
                        })
                    except (ValueError, IndexError):
                        pass
            
            # Calculate RDA
            if age < 19:
                if gender == 'female':
                    rda = {'energy_kcal': 2000, 'protein_g': 46, 'iron_mg': 15, 'calcium_mg': 1300}
                else:
                    rda = {'energy_kcal': 2400, 'protein_g': 52, 'iron_mg': 11, 'calcium_mg': 1300}
            else:
                if gender == 'female':
                    rda = {'energy_kcal': 2100, 'protein_g': 46, 'iron_mg': 29, 'calcium_mg': 1000}
                else:
                    rda = {'energy_kcal': 2500, 'protein_g': 56, 'iron_mg': 11, 'calcium_mg': 1000}
            
            total_fluid_l = total_fluid_ml / 1000
            ai_liters = 3.7 if gender == 'male' else 2.7
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Nutrition Summary"
            
            ws['A1'] = "24-HOUR DIETARY RECALL REPORT"
            ws['A1'].font = Font(bold=True, size=16)
            
            ws['A3'] = "Patient Information"
            ws['A3'].font = Font(bold=True, size=12, color="2c5e2e")
            ws['A4'] = "Name:"
            ws['B4'] = name or "Not provided"
            ws['A5'] = "Age:"
            ws['B5'] = age
            ws['A6'] = "Gender:"
            ws['B6'] = "Female" if gender == 'female' else "Male"
            ws['A7'] = "Date:"
            ws['B7'] = datetime.now().strftime('%Y-%m-%d %H:%M')
            
            ws['A9'] = "Total Nutrients"
            ws['A9'].font = Font(bold=True, size=12, color="2c5e2e")
            
            nutrients = [
                ('Energy (kcal)', total_energy, rda['energy_kcal']),
                ('Protein (g)', total_protein, rda['protein_g']),
                ('Iron (mg)', total_iron, rda['iron_mg']),
                ('Calcium (mg)', total_calcium, rda['calcium_mg']),
            ]
            
            row = 10
            for nutrient_name, value, target in nutrients:
                ws[f'A{row}'] = nutrient_name
                ws[f'B{row}'] = f"{value:.1f}"
                ws[f'C{row}'] = f"Target: {target}"
                percent = (value / target) * 100 if target > 0 else 0
                ws[f'D{row}'] = f"{percent:.0f}%"
                if percent < 70:
                    ws[f'D{row}'].font = Font(color="dc3545", bold=True)
                row += 1
            
            ws[f'A{row+2}'] = "Hydration"
            ws[f'A{row+2}'].font = Font(bold=True, size=12, color="2c5e2e")
            ws[f'A{row+3}'] = "Total Fluids:"
            ws[f'B{row+3}'] = f"{total_fluid_ml:.0f} ml ({total_fluid_l:.1f} L)"
            ws[f'A{row+4}'] = "Daily Target:"
            ws[f'B{row+4}'] = f"{ai_liters} L"
            
            # Auto-adjust columns
            for col in ['A', 'B', 'C', 'D']:
                max_length = 0
                for row_num in range(1, row + 10):
                    cell_value = ws[f'{col}{row_num}'].value
                    if cell_value and len(str(cell_value)) > max_length:
                        max_length = len(str(cell_value))
                adjusted_width = min(max_length + 2, 40)
                ws.column_dimensions[col].width = adjusted_width
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            filename = f"nutrition_recall_{name.replace(' ', '_') if name else 'report'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            return response
            
        else:
            return HttpResponse("Invalid request method", status=400)
            
    except Exception as e:
        return HttpResponse(f"Error exporting recall data: {str(e)}", status=500)

def nutrient_calculator(request):
    foods_data = get_all_foods_supabase()
    foods = dict_list_to_food_objects(foods_data)
    categories = Category.objects.all()
    
    result = None
    food_selected = None
    amount = 100
    unit = 'grams'
    gender = 'female'
    age = 30
    available_units = []
    rda = None
    
    pre_selected_food = request.GET.get('food', '')
    if pre_selected_food and not request.POST:
        try:
            food_selected_data = search_foods_supabase(pre_selected_food)
            if food_selected_data:
                food_selected = dict_to_food_object(food_selected_data[0])
                available_units = UnitConversion.objects.filter(food_id=food_selected.id)
        except:
            pass
    
    if request.method == 'POST':
        food_id = request.POST.get('food_id')
        
        if not food_id:
            gender = request.POST.get('gender', 'female')
            try:
                age = int(request.POST.get('age', 30))
            except ValueError:
                age = 30
            return render(request, 'foods/calculator.html', {
                'foods': foods,
                'categories': categories,
                'food_selected': None,
                'amount': 100,
                'unit': 'grams',
                'gender': gender,
                'age': age,
                'available_units': [],
                'result': None,
                'rda': None,
            })
        
        amount = float(request.POST.get('amount', 100))
        unit = request.POST.get('unit', 'grams')
        gender = request.POST.get('gender', 'female')
        age = int(request.POST.get('age', 30))
        
        try:
            food_selected_data = get_food_by_id_supabase(int(food_id))
            if food_selected_data:
                food_selected = dict_to_food_object(food_selected_data)
                available_units = UnitConversion.objects.filter(food_id=food_id)
                
                grams = amount
                if unit != 'grams':
                    try:
                        conversion = UnitConversion.objects.get(food_id=food_id, unit_name=unit)
                        grams = amount * conversion.grams
                    except:
                        grams = amount
                
                result = {
                    'food_name': food_selected.food_name,
                    'amount': amount,
                    'unit': unit,
                    'grams': grams,
                    'energy_kcal': (grams / 100) * food_selected.energy_kcal,
                    'protein_g': (grams / 100) * food_selected.protein_g,
                    'fat_g': (grams / 100) * food_selected.fat_g,
                    'carbohydrate_g': (grams / 100) * food_selected.carbohydrate_g,
                    'fiber_g': (grams / 100) * food_selected.fiber_g,
                    'iron_mg': (grams / 100) * food_selected.iron_mg,
                    'calcium_mg': (grams / 100) * food_selected.calcium_mg,
                }
                
                if age < 19:
                    if gender == 'female':
                        rda = {'energy_kcal': 2000, 'protein_g': 46, 'iron_mg': 15, 'calcium_mg': 1300}
                    else:
                        rda = {'energy_kcal': 2400, 'protein_g': 52, 'iron_mg': 11, 'calcium_mg': 1300}
                else:
                    if gender == 'female':
                        rda = {'energy_kcal': 2100, 'protein_g': 46, 'iron_mg': 29, 'calcium_mg': 1000}
                    else:
                        rda = {'energy_kcal': 2500, 'protein_g': 56, 'iron_mg': 11, 'calcium_mg': 1000}
                
                result['energy_percent'] = (result['energy_kcal'] / rda['energy_kcal']) * 100
                result['iron_percent'] = (result['iron_mg'] / rda['iron_mg']) * 100 if rda['iron_mg'] > 0 else 0
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'result': result,
                        'rda': rda,
                    })
            
        except Exception as e:
            print(f"Calculator error: {e}")
            pass
    
    return render(request, 'foods/calculator.html', {
        'foods': foods,
        'categories': categories,
        'food_selected': food_selected,
        'amount': amount,
        'unit': unit,
        'gender': gender,
        'age': age,
        'available_units': available_units,
        'result': result,
        'rda': rda,
    })

def recall_24hr(request):
    """24-hour dietary recall with searchable foods and fluids"""
    foods_data = get_all_foods_supabase()
    foods = dict_list_to_food_objects(foods_data)
    
    meals = [
        {'id': 1, 'name': 'Breakfast'},
        {'id': 2, 'name': 'Morning Snack'},
        {'id': 3, 'name': 'Lunch'},
        {'id': 4, 'name': 'Afternoon Snack'},
        {'id': 5, 'name': 'Dinner'},
        {'id': 6, 'name': 'Evening Snack'},
    ]
    
    results = None
    
    if request.method == 'POST':
        name = request.POST.get('name', '')
        age = request.POST.get('age', '')
        gender = request.POST.get('gender', 'female')
        
        total_energy = 0
        total_protein = 0
        total_iron = 0
        total_fiber = 0
        total_calcium = 0
        total_vitamin_a = 0
        total_fluid_ml = 0
        
        try:
            age = int(age) if age else 30
        except ValueError:
            age = 30
        
        fluid_ids = request.POST.getlist('fluid_id[]')
        fluid_amounts = request.POST.getlist('fluid_amount[]')
        fluid_units = request.POST.getlist('fluid_unit[]')
        
        for i in range(len(fluid_ids)):
            if fluid_ids[i] and fluid_amounts[i]:
                try:
                    amount = float(fluid_amounts[i])
                    unit = fluid_units[i]
                    
                    if unit == 'ml':
                        total_fluid_ml += amount
                    elif unit == 'cup':
                        total_fluid_ml += amount * 240
                    elif unit == 'glass':
                        total_fluid_ml += amount * 250
                    elif unit == 'bottle':
                        total_fluid_ml += amount * 500
                except (ValueError, IndexError):
                    pass
        
        for meal in meals:
            meal_id = meal['id']
            food_ids = request.POST.getlist(f'food_id_{meal_id}[]')
            amounts = request.POST.getlist(f'amount_{meal_id}[]')
            units = request.POST.getlist(f'unit_{meal_id}[]')
            
            for i in range(len(food_ids)):
                if food_ids[i] and amounts[i]:
                    try:
                        food_data = get_food_by_id_supabase(int(food_ids[i]))
                        if food_data:
                            food = dict_to_food_object(food_data)
                            amount = float(amounts[i])
                            unit = units[i] if i < len(units) else 'grams'
                            
                            grams = amount
                            if unit != 'grams':
                                try:
                                    conversion = UnitConversion.objects.get(food_id=food_ids[i], unit_name=unit)
                                    grams = amount * conversion.grams
                                except:
                                    grams = amount
                            
                            total_energy += (grams / 100) * food.energy_kcal
                            total_protein += (grams / 100) * food.protein_g
                            total_iron += (grams / 100) * food.iron_mg
                            total_fiber += (grams / 100) * food.fiber_g
                            total_calcium += (grams / 100) * food.calcium_mg
                            total_vitamin_a += (grams / 100) * getattr(food, 'vitamin_a_rae_ug', 0)
                            total_fluid_ml += (grams / 100) * getattr(food, 'water_g', 0)
                            
                    except Exception as e:
                        print(f"Recall error: {e}")
                        pass
        
        if age < 4:
            ai_liters = 1.3
        elif age < 9:
            ai_liters = 1.7
        elif age < 14:
            ai_liters = 2.4 if gender == 'male' else 2.1
        elif age < 19:
            ai_liters = 3.3 if gender == 'male' else 2.3
        else:
            ai_liters = 3.7 if gender == 'male' else 2.7
        
        total_fluid_l = total_fluid_ml / 1000
        fluid_percent = (total_fluid_l / ai_liters) * 100 if ai_liters > 0 else 0
        
        results = {
            'total_energy': total_energy,
            'total_protein': total_protein,
            'total_iron': total_iron,
            'total_fiber': total_fiber,
            'total_calcium': total_calcium,
            'total_vitamin_a': total_vitamin_a,
            'total_fluid_ml': total_fluid_ml,
            'total_fluid_l': total_fluid_l,
            'ai_liters': ai_liters,
            'fluid_percent': fluid_percent,
            'name': name,
            'age': age,
            'gender': gender,
        }
        
        if fluid_percent < 80:
            results['hydration_message'] = '⚠️ You might be running low on fluids. Try to drink more water throughout the day.'
        elif fluid_percent <= 120:
            results['hydration_message'] = '✅ Great job! Your hydration is on point.'
        else:
            results['hydration_message'] = '💧 You\'re well hydrated!'
        
        # RDI analysis
        if gender == 'female':
            if age < 19:
                rdi = {'energy_kcal': 2000, 'protein_g': 46, 'iron_mg': 15, 'calcium_mg': 1300}
            else:
                rdi = {'energy_kcal': 2100, 'protein_g': 46, 'iron_mg': 29, 'calcium_mg': 1000}
        else:
            if age < 19:
                rdi = {'energy_kcal': 2400, 'protein_g': 52, 'iron_mg': 11, 'calcium_mg': 1300}
            else:
                rdi = {'energy_kcal': 2500, 'protein_g': 56, 'iron_mg': 11, 'calcium_mg': 1000}
        
        results['rdi'] = rdi
        results['rdi_percentages'] = {
            'energy_percent': (total_energy / rdi['energy_kcal']) * 100,
            'protein_percent': (total_protein / rdi['protein_g']) * 100,
            'iron_percent': (total_iron / rdi['iron_mg']) * 100,
            'calcium_percent': (total_calcium / rdi['calcium_mg']) * 100,
        }
        
        results['status_messages'] = []
        
        if results['rdi_percentages']['iron_percent'] < 70:
            results['status_messages'].append({
                'nutrient': 'Iron',
                'status': 'low',
                'message': f"⚠️ Your iron intake ({total_iron:.1f}mg) is only {results['rdi_percentages']['iron_percent']:.0f}% of daily needs.",
                'suggestions': ['sukuma wiki', 'beans', 'dagaa omena']
            })
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'results': results,
                'name': name or 'there',
            })
    
    return render(request, 'foods/recall_24hr.html', {
        'foods': foods,
        'meals': meals,
        'results': results,
    })

def compare_foods(request):
    foods_data = get_all_foods_supabase()
    foods = dict_list_to_food_objects(foods_data)
    
    food1 = None
    food2 = None
    comparison = None
    messages = []
    
    if request.method == 'POST':
        food1_id = request.POST.get('food1')
        food2_id = request.POST.get('food2')
        
        if food1_id and food2_id:
            try:
                food1_data = get_food_by_id_supabase(int(food1_id))
                food2_data = get_food_by_id_supabase(int(food2_id))
                
                if food1_data and food2_data:
                    food1 = dict_to_food_object(food1_data)
                    food2 = dict_to_food_object(food2_data)
                    
                    nutrients = [
                        {'name': 'Energy (kcal)', 'key': 'energy_kcal', 'unit': 'kcal', 'higher_is': 'better'},
                        {'name': 'Protein (g)', 'key': 'protein_g', 'unit': 'g', 'higher_is': 'better'},
                        {'name': 'Fiber (g)', 'key': 'fiber_g', 'unit': 'g', 'higher_is': 'better'},
                        {'name': 'Iron (mg)', 'key': 'iron_mg', 'unit': 'mg', 'higher_is': 'better'},
                        {'name': 'Calcium (mg)', 'key': 'calcium_mg', 'unit': 'mg', 'higher_is': 'better'},
                    ]
                    
                    comparison = []
                    for n in nutrients:
                        val1 = getattr(food1, n['key'], 0)
                        val2 = getattr(food2, n['key'], 0)
                        
                        if n['higher_is'] == 'better':
                            if val1 > val2:
                                winner = 1
                            elif val2 > val1:
                                winner = 2
                            else:
                                winner = 0
                        else:
                            winner = 0
                        
                        max_val = max(val1, val2)
                        if max_val > 0:
                            pct1 = (val1 / max_val) * 100
                            pct2 = (val2 / max_val) * 100
                        else:
                            pct1 = 0
                            pct2 = 0
                        
                        comparison.append({
                            'name': n['name'],
                            'key': n['key'],
                            'unit': n['unit'],
                            'val1': val1,
                            'val2': val2,
                            'pct1': pct1,
                            'pct2': pct2,
                            'winner': winner,
                        })
                    
                    if food1.iron_mg > food2.iron_mg * 1.5:
                        messages.append(f"🔴 {food1.food_name[:30]} has {food1.iron_mg:.1f}mg iron — more than {food2.food_name[:30]}")
                    elif food2.iron_mg > food1.iron_mg * 1.5:
                        messages.append(f"🔴 {food2.food_name[:30]} has {food2.iron_mg:.1f}mg iron — more than {food1.food_name[:30]}")
                    
                    if not messages:
                        messages.append("💡 These foods have similar nutritional profiles.")
                    
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({
                            'success': True,
                            'food1': {'name': food1.food_name, 'id': food1.id},
                            'food2': {'name': food2.food_name, 'id': food2.id},
                            'comparison': comparison,
                            'messages': messages,
                        })
                
            except Exception as e:
                print(f"Compare error: {e}")
                pass
    
    return render(request, 'foods/compare.html', {
        'foods': foods,
        'food1': food1,
        'food2': food2,
        'comparison': comparison,
        'messages': messages,
    })

# ========== API ENDPOINTS ==========
def api_foods(request):
    foods_data = get_all_foods_supabase()
    # Limit to 50 for performance
    foods_data = foods_data[:50]
    # Format for API response
    result = []
    for food in foods_data:
        result.append({
            'id': food.get('id'),
            'food_name': food.get('food_name'),
            'category_name': food.get('category'),
            'energy_kcal': food.get('energy_kcal'),
            'protein_g': food.get('protein_g'),
            'fat_g': food.get('fat_g'),
            'carbohydrate_g': food.get('carbohydrate_g'),
            'fiber_g': food.get('fiber_g'),
            'iron_mg': food.get('iron_mg'),
            'calcium_mg': food.get('calcium_mg'),
        })
    return JsonResponse(result, safe=False)

def api_food_detail(request, food_id):
    try:
        food_data = get_food_by_id_supabase(food_id)
        if food_data:
            return JsonResponse({
                'id': food_data.get('id'),
                'name': food_data.get('food_name'),
                'category': food_data.get('category'),
                'energy_kcal': food_data.get('energy_kcal'),
                'protein_g': food_data.get('protein_g'),
                'fat_g': food_data.get('fat_g'),
                'carbohydrate_g': food_data.get('carbohydrate_g'),
                'fiber_g': food_data.get('fiber_g'),
                'iron_mg': food_data.get('iron_mg'),
                'calcium_mg': food_data.get('calcium_mg'),
            })
        else:
            return JsonResponse({'error': 'Food not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=404)

def api_search(request):
    query = request.GET.get('q', '')
    if query:
        foods_data = search_foods_supabase(query)
        result = []
        for food in foods_data:
            result.append({
                'id': food.get('id'),
                'food_name': food.get('food_name'),
                'category_name': food.get('category'),
                'energy_kcal': food.get('energy_kcal'),
                'protein_g': food.get('protein_g'),
                'iron_mg': food.get('iron_mg'),
            })
        return JsonResponse(result, safe=False)
    return JsonResponse([], safe=False)

def api_categories(request):
    categories = Category.objects.annotate(food_count=Count('foods')).values('id', 'name', 'food_count')
    return JsonResponse(list(categories), safe=False)

def create_admin(request):
    username = 'admin123'
    email = 'admin@example.com'
    password = 'admin123'
    
    if not User.objects.filter(username=username).exists():
        User.objects.create_superuser(username=username, email=email, password=password)
        return HttpResponse(f"✅ Admin user created! Username: {username}, Password: {password}")
    else:
        return HttpResponse(f"⚠️ User '{username}' already exists. Try logging in.")

@csrf_exempt
def health_check(request):
    """Health check endpoint"""
    return JsonResponse({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'message': 'Django server is running with Supabase'
    })